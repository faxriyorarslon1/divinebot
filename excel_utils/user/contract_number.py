from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl.styles import Font

from api.orders.contract_number import ContractNumber
from configs.constants import BASE_URL
from os.path import join as join_path

from excel_utils import check_excel
from excel_utils.product_report import CONTRACT_NUMBER_EXCEL_PATH

ORDER_EXCEL_PATH = join_path(BASE_URL, 'excel_utils', 'excel', 'contract_number')

df = pd.DataFrame(columns=["Id", 'FIO', 'Viloyati', 'INN', 'Comment', 'Umumiy Summa', 'To\'lov Turi'])

year = datetime.now().year
order_path = f"contract_number_{year}.xlsx"


def create_contract_excel(_id, time, company_name, inn, phone_number, village, first_name, last_name=None):
    check_excel_params = check_excel(ORDER_EXCEL_PATH, order_path)
    if check_excel_params.__eq__("file yoq"):
        book1 = openpyxl.Workbook()
        sheet1 = book1.active
        columns = ["Shartnoma Raqam", 'Vaqt', 'Korxona Nomi', 'INN', 'Telefon Raqam', 'Viloyat', 'Tibbiy Vakil']
        sheet1['A1'] = columns[0]
        sheet1["B1"] = columns[1]
        sheet1["C1"] = columns[2]
        sheet1["D1"] = columns[3]
        sheet1["E1"] = columns[4]
        sheet1["F1"] = columns[5]
        sheet1["G1"] = columns[6]
        sheet1.column_dimensions['A'].width = 30
        sheet1.column_dimensions['B'].width = 20
        sheet1.column_dimensions['C'].width = 30
        sheet1.column_dimensions['D'].width = 30
        sheet1.column_dimensions['E'].width = 30
        sheet1.column_dimensions['F'].width = 30
        sheet1.column_dimensions['G'].width = 30
        if last_name:
            row = (_id, time, company_name, inn, phone_number, village, f"{first_name} {last_name}")
        else:
            row = (_id, time, company_name, inn, phone_number, village, f"{first_name}")
        sheet1.append(row)
        book1.save(f'{join_path(ORDER_EXCEL_PATH, order_path)}')
        book = openpyxl.load_workbook(f'{join_path(ORDER_EXCEL_PATH, order_path)}')
        book.save(f'{join_path(ORDER_EXCEL_PATH, order_path)}')
    else:
        book = openpyxl.load_workbook(f'{join_path(ORDER_EXCEL_PATH, order_path)}')
        sheet = book.active
        sheet.cell(f"C{sheet.max_row}").style.aligment.wrap_text = True
        if last_name:
            row = (_id, time, company_name, inn, phone_number, village, f"{first_name} {last_name}")
        else:
            row = (_id, time, company_name, inn, phone_number, village, f"{first_name}")
        sheet.append(row)
        book.save(f'{join_path(ORDER_EXCEL_PATH, order_path)}')


def write_contract_number_agent(_id, time, company_name, inn, phone_number, village, first_name, last_name=None,
                                token=None):
    book = openpyxl.load_workbook(f'{join_path(ORDER_EXCEL_PATH, order_path)}')
    sheet = book.active
    contract_number = ContractNumber(token=token).get_excel()
    check_excel_params = check_excel(CONTRACT_NUMBER_EXCEL_PATH, contract_number['file_name'])
    if not check_excel_params.__eq__("file yoq"):
        if last_name:
            row = (_id, time, company_name, inn, phone_number, village, f"{first_name} {last_name}")
        else:
            row = (_id, time, company_name, inn, phone_number, village, f"{first_name}")
        sheet.append(row)
        book.save(f'{join_path(CONTRACT_NUMBER_EXCEL_PATH, contract_number["file_name"])}')
        return True
    else:
        return False


def get_contract_number(inn, token):
    data = list()
    c_n = ContractNumber(token=token).get_excel()
    EXCEL_PATH_VIZIT = join_path(BASE_URL, 'excel_utils', 'excel', 'contract_number')
    excel_path = join_path(EXCEL_PATH_VIZIT, c_n['file_name'])
    wb_obj = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb_obj.active
    max_row = sheet.max_row
    for i in range(2, max_row + 1):
        data_income = {}
        excel_inn = str(sheet.cell(row=i, column=4).value).replace(" ", '')
        try:
            excel_inn = int(excel_inn)
        except Exception:
            print("Raise")
        if inn == excel_inn:
            data_income['inn'] = sheet.cell(row=i, column=4).value
            data_income['number'] = sheet.cell(row=i, column=1).value
            data_income['time'] = sheet.cell(row=i, column=2).value
            data_income['company_name'] = sheet.cell(row=i, column=3).value
            data_income['phone_number'] = sheet.cell(row=i, column=5).value
            data_income['region'] = sheet.cell(row=i, column=7).value
            data_income['mp'] = sheet.cell(row=i, column=8).value
            data.append(data_income)
    return data
