from datetime import datetime
from os.path import join as join_path

import openpyxl

from configs.constants import BASE_URL
from excel_utils import check_excel

from utils.geolocation_ import location_address

LOCATION_EXCEL_PATH = join_path(BASE_URL, 'excel_utils', 'excel', 'location')

year = datetime.now().year
month = datetime.now().month


def location_excel(created_at, created_by: str, latitude, longitude, village):
    created_by_for_excel = created_by.replace(" ", "_")
    excel_path = f"{created_by_for_excel}_location_{year}_{month}.xlsx"
    location = location_address(latitude, longitude)
    check_excel_params = check_excel(LOCATION_EXCEL_PATH, excel_path)
    if check_excel_params.__eq__("file yoq"):
        book1 = openpyxl.Workbook()
        sheet1 = book1.active
        columns = ["Yaratilgan Vaqt", 'Kim Tomonidan', "Joylashuv", "Viloyat"]
        sheet1['A1'] = columns[0]
        sheet1["B1"] = columns[1]
        sheet1["C1"] = columns[2]
        sheet1["D1"] = columns[3]
        book1.save(join_path(LOCATION_EXCEL_PATH, excel_path))
        book = openpyxl.load_workbook(join_path(LOCATION_EXCEL_PATH, excel_path))
        sheet = book.active
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 60
        sheet.column_dimensions['D'].width = 30
        row = (created_at, created_by, location, village)
        sheet.append(row)
        book.save(join_path(LOCATION_EXCEL_PATH, excel_path))
    else:
        book = openpyxl.load_workbook(join_path(LOCATION_EXCEL_PATH, excel_path))
        sheet = book.active
        row = (created_at, created_by, location, village)
        sheet.append(row)
        book.save(join_path(LOCATION_EXCEL_PATH, excel_path))
