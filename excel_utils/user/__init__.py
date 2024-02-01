import datetime
import openpyxl
from configs.constants import BASE_URL
from excel_utils import check_excel
from os.path import join as join_path

USER_EXCEL_PATH = join_path(BASE_URL, 'excel_utils', 'excel', 'user')
VIZIT_EXCEL_PATH = join_path(BASE_URL, 'excel_utils', 'excel', 'vizit')


def create_doctor_excel(created_at: object, created_by: object, phone_number: object, village: object, city: object,
                        lpu: object, doctor_name: object, doctor_phone: object,
                        comment: object) -> object:
    created_by_for_excel = created_by.replace(" ", "_")
    year = datetime.datetime.now().year
    month = datetime.datetime.now().month
    excel_path = f"{created_by_for_excel}vizit_excel{year}_{month}.xlsx"
    check_excel_params = check_excel(USER_EXCEL_PATH, excel_path)
    if check_excel_params.__eq__("file yoq"):
        book1 = openpyxl.Workbook()
        sheet1 = book1.active
        columns = ["Vaqti", 'Kimdan', 'Telefon Nomer', "Viloyat", 'Shaxar', 'LPU', 'Doktor ismi', "Doktor telefoni",
                   'Izoh']
        sheet1['A1'] = columns[0]
        sheet1["B1"] = columns[1]
        sheet1["C1"] = columns[2]
        sheet1["D1"] = columns[3]
        sheet1["E1"] = columns[4]
        sheet1["F1"] = columns[5]
        sheet1["G1"] = columns[6]
        sheet1['H1'] = columns[7]
        sheet1['I1'] = columns[8]
        book1.save(f'{join_path(USER_EXCEL_PATH, excel_path)}')
        book = openpyxl.load_workbook(f'{join_path(USER_EXCEL_PATH, excel_path)}')
        sheet = book.active
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 30
        row = (created_at, created_by, phone_number, village, city, lpu, doctor_name, doctor_phone, comment)
        sheet.append(row)
        book.save(f'{join_path(USER_EXCEL_PATH, excel_path)}')
    else:
        book = openpyxl.load_workbook(f'{join_path(USER_EXCEL_PATH, excel_path)}')
        sheet = book.active
        row = (created_at, created_by, phone_number, village, city, lpu, doctor_name, doctor_phone, comment)
        sheet.append(row)
        book.save(f'{join_path(USER_EXCEL_PATH, excel_path)}')
