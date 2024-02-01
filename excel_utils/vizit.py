from datetime import datetime
from os.path import join as join_path

import openpyxl

from api.users import district_get_all, district_retrieve
from configs.constants import BASE_URL
from excel_utils import check_excel

VIZIT_EXCEL_PATH = join_path(BASE_URL, 'excel_utils', 'excel', 'vizit')


def get_all_user_excel():
    district = district_get_all()
    for d in district['results']:
        for i in range(1, 13):
            excel_path = f"obshi_vizit_excel_2024_{i}_{d['name']}.xlsx"
            check_excel_params = check_excel(VIZIT_EXCEL_PATH, excel_path)
            if check_excel_params.__eq__("file yoq"):
                book = openpyxl.Workbook()
                sheet1 = book.active
                columns = ["Vaqti", 'Kimdan', 'Telefon Nomer', "Viloyat", 'Shaxar', 'LPU', 'Doktor ismi',
                           'Mutaxasisligi',
                           "Kategoriyasi",
                           "Doktor telefoni",
                           'Izoh']
                sheet1['A1'] = columns[0]
                sheet1["B1"] = columns[1]
                sheet1["C1"] = columns[2]
                sheet1["D1"] = columns[3]
                sheet1["E1"] = columns[4]
                sheet1["F1"] = columns[5]
                sheet1["G1"] = columns[6]
                sheet1['H1'] = columns[7]
                sheet1['I1'] = columns[8]
                sheet1['J1'] = columns[9]
                sheet1['K1'] = columns[10]
                # book1.save(f'{join_path(VIZIT_EXCEL_PATH, excel_path)}')
                book.save(f'{join_path(VIZIT_EXCEL_PATH, excel_path)}')


#
# print(get_all_user_excel())


def create_doctor_all_excel(created_at: object, created_by: object, phone_number: object, village: object, city: object,
                            lpu: object, doctor_name: object, category, d_type, doctor_phone: object,
                            comment: object) -> object:
    excel_path = f"obshi_vizit_excel_{village}.xlsx"
    check_excel_params = check_excel(VIZIT_EXCEL_PATH, excel_path)
    if check_excel_params.__eq__("file yoq"):
        book1 = openpyxl.Workbook()
        sheet1 = book1.active
        columns = ["Vaqti", 'Kimdan', 'Telefon Nomer', "Viloyat", 'Shaxar', 'LPU', 'Doktor ismi', 'Mutaxasisligi',
                   "Kategoriyasi",
                   "Doktor telefoni",
                   'Izoh']
        sheet1['A1'] = columns[0]
        sheet1["B1"] = columns[1]
        sheet1["C1"] = columns[2]
        sheet1["D1"] = columns[3]
        sheet1["E1"] = columns[4]
        sheet1["F1"] = columns[5]
        sheet1["G1"] = columns[6]
        sheet1['H1'] = columns[7]
        sheet1['I1'] = columns[8]
        sheet1['J1'] = columns[9]
        sheet1['K1'] = columns[10]
        book1.save(f'{join_path(VIZIT_EXCEL_PATH, excel_path)}')
        book = openpyxl.load_workbook(f'{join_path(VIZIT_EXCEL_PATH, excel_path)}')
        sheet = book.active
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 30
        sheet.column_dimensions['H'].width = 30
        sheet.column_dimensions['I'].width = 30
        sheet.column_dimensions['J'].width = 30
        sheet.column_dimensions['K'].width = 60
        row = (
            created_at, created_by, phone_number, village, city, lpu, doctor_name, category, d_type, doctor_phone,
            comment)
        sheet.append(row)
        book.save(f'{join_path(VIZIT_EXCEL_PATH, excel_path)}')
    else:
        book = openpyxl.load_workbook(f'{join_path(VIZIT_EXCEL_PATH, excel_path)}')
        sheet = book.active
        row = (
            created_at, created_by, phone_number, village, city, lpu, doctor_name, category, d_type, doctor_phone,
            comment)
        sheet.append(row)
        book.save(f'{join_path(VIZIT_EXCEL_PATH, excel_path)}')


def create_mp_doctor_or_pharmacy_all_excel(created_at,
                                           created_by,
                                           village,
                                           city,
                                           lpu,
                                           doctor_name,
                                           category,
                                           d_type,
                                           doctor_phone,
                                           comment,
                                           pharmacy_name,
                                           pharmacy_address,
                                           mp=None,
                                           preparation=1,
                                           communication=1,
                                           the_need=1,
                                           presentation=1,
                                           protest=1,
                                           agreement=1,
                                           analysis=1,

                                           ) -> None:
    # created_by_for_excel = created_by.replace(" ", "_")
    VIZIT_EXCEL_PATH = join_path(BASE_URL, 'excel_utils', 'excel', 'double_excel', 'doctor')
    if preparation != "Ma'lumot yo'q":
        urtacha = (int(presentation) + int(preparation) + int(communication) + int(the_need) + int(protest) + int(
            agreement) + int(analysis)) / 7
    else:
        urtacha = "Ma'lumot yo'q"
    excel_path = "pharmacy_or_vizit.xlsx"
    year = created_at[:11]
    hour = created_at[11:19]
    check_excel_params = check_excel(VIZIT_EXCEL_PATH, excel_path)
    if check_excel_params.__eq__("file yoq"):
        book1 = openpyxl.Workbook()
        sheet1 = book1.active
        columns = ["Vaqti(Yil-Oy-Kun)", "Vaqti(Soat-Minut)", 'Kimdan', "Viloyat",
                   'Shaxar', 'LPU', 'Doktor ismi',
                   'Mutaxasisligi', "Kategoriyasi",
                   "Doktor telefoni", 'Izoh', "Dorixona Nomi",
                   "Dorixona Manzili", "MP",
                   "Tayyorgarlik", 'Muloqot', 'Extiyoj',
                   'Taqdimot', "E'tiroz",
                   'Kelishuv', 'Taxlil', "O'rtacha"]
        sheet1['A1'] = columns[0]
        sheet1["B1"] = columns[1]
        sheet1["C1"] = columns[2]
        sheet1["D1"] = columns[3]
        sheet1["E1"] = columns[4]
        sheet1["F1"] = columns[5]
        sheet1["G1"] = columns[6]
        sheet1['H1'] = columns[7]
        sheet1['I1'] = columns[8]
        sheet1['J1'] = columns[9]
        sheet1['K1'] = columns[10]
        sheet1['L1'] = columns[11]
        sheet1['M1'] = columns[12]
        sheet1['N1'] = columns[13]
        sheet1['O1'] = columns[14]
        sheet1['P1'] = columns[15]
        sheet1['Q1'] = columns[16]
        sheet1['R1'] = columns[17]
        sheet1['S1'] = columns[18]
        sheet1['T1'] = columns[19]
        sheet1['U1'] = columns[20]
        sheet1['V1'] = columns[21]
        book1.save(f'{join_path(VIZIT_EXCEL_PATH, excel_path)}')
        book = openpyxl.load_workbook(f'{join_path(VIZIT_EXCEL_PATH, excel_path)}')
        sheet = book.active
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 30
        sheet.column_dimensions['H'].width = 30
        sheet.column_dimensions['I'].width = 30
        sheet.column_dimensions['J'].width = 30
        sheet.column_dimensions['K'].width = 60
        sheet.column_dimensions['L'].width = 30
        sheet.column_dimensions['M'].width = 30
        sheet.column_dimensions['N'].width = 30
        sheet.column_dimensions['O'].width = 30
        sheet.column_dimensions['P'].width = 30
        sheet.column_dimensions['Q'].width = 30
        sheet.column_dimensions['R'].width = 30
        sheet.column_dimensions['S'].width = 30
        sheet.column_dimensions['T'].width = 30
        sheet.column_dimensions['U'].width = 30
        sheet.column_dimensions['V'].width = 30
        row = (
            year, hour, created_by, village, city, lpu, doctor_name, category, d_type, doctor_phone,
            comment, pharmacy_name, pharmacy_address, mp, preparation, communication, the_need, presentation, protest,
            agreement, analysis, urtacha)
        sheet.append(row)
        book.save(f'{join_path(VIZIT_EXCEL_PATH, excel_path)}')
    else:
        book = openpyxl.load_workbook(f'{join_path(VIZIT_EXCEL_PATH, excel_path)}')
        sheet = book.active
        row = (
            year, hour, created_by, village, city, lpu, doctor_name, category, d_type, doctor_phone,
            comment, pharmacy_name, pharmacy_address, mp, preparation, communication, the_need, presentation, protest,
            agreement, analysis, urtacha)
        sheet.append(row)
        book.save(f'{join_path(VIZIT_EXCEL_PATH, excel_path)}')


print(str(datetime.now())[:11])
print(str(datetime.now())[11:19])
