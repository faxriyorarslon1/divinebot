import json
import openpyxl
import pandas as pd
from datetime import datetime
from os.path import join as join_path
from openpyxl.styles import Font

from Tranlate.tranlate_config import translate_cyrillic_or_latin
from configs.constants import BASE_URL
from excel_utils import check_excel
from excel_utils.product_report import PRODUCT_EXCEL_PATH, INCOME_EXCEL_PATH, WRONG_FILE, DEBIT_EXCEL_PATH

ORDER_EXCEL_PATH = join_path(BASE_URL, 'excel_utils', 'excel', 'order')

df = pd.DataFrame(columns=["Id", 'FIO', 'Viloyati', 'INN', 'Comment', 'Umumiy Summa', 'To\'lov Turi'])

year = datetime.now().year
month = datetime.now().month
order_path = f"{year}{month}order.xlsx"


def json_read_util():
    with open('excel_count.json', 'r') as file:
        files = json.load(file)
        return files['order_count']


def json_write_util():
    files = json_read_util()
    files += 1
    with open('excel_count.json', 'w') as file:
        json.dump({"order_count": files}, file)


# print(json_write_util())


def create_order_excel(_id, first_name, last_name, city, inn, comment, total_price, type_price):
    check_excel_params = check_excel(ORDER_EXCEL_PATH, order_path)
    # json_count = json_read_util()
    if check_excel_params.__eq__("file yoq"):
        with open('excel_count.json', 'w') as file:
            json.dump({"order_count": 2}, file)
        book1 = openpyxl.Workbook()
        sheet1 = book1.active
        columns = ["Id", 'FIO', 'Viloyati', 'INN', 'Comment', 'Umumiy Summa', 'To\'lov Turi']
        sheet1['A1'] = columns[0]
        sheet1["B1"] = columns[1]
        sheet1["C1"] = columns[2]
        sheet1["D1"] = columns[3]
        sheet1["E1"] = columns[4]
        sheet1["F1"] = columns[5]
        sheet1["G1"] = columns[6]
        sheet1.cell(row=1, column=1).font = Font(size=20, name='Times New Roman')
        sheet1.cell(row=1, column=2).font = Font(size=20, name='Times New Roman')
        sheet1.cell(row=1, column=3).font = Font(size=20, name='Times New Roman')
        sheet1.cell(row=1, column=4).font = Font(size=20, name='Times New Roman')
        sheet1.cell(row=1, column=5).font = Font(size=20, name='Times New Roman')
        sheet1.cell(row=1, column=6).font = Font(size=20, name='Times New Roman')
        sheet1.cell(row=1, column=7).font = Font(size=20, name='Times New Roman')
        book1.save(f'{join_path(ORDER_EXCEL_PATH, order_path)}')
        book = openpyxl.load_workbook(f'{join_path(ORDER_EXCEL_PATH, order_path)}')
        book.save(f'{join_path(ORDER_EXCEL_PATH, order_path)}')
    else:
        book = openpyxl.load_workbook(f'{join_path(ORDER_EXCEL_PATH, order_path)}')
        sheet = book.active
        sheet.cell(f"C{sheet.max_row}").style.aligment.wrap_text = True
        row = (_id, f"{first_name} {last_name}", city, inn, comment, total_price, type_price)
        sheet.append(row)
        book.save(f'{join_path(ORDER_EXCEL_PATH, order_path)}')


def add_excel_product(path):
    check_excel_params = check_excel(PRODUCT_EXCEL_PATH, path)
    if not check_excel_params.__eq__("file yoq"):
        excel_path = join_path(PRODUCT_EXCEL_PATH, path)
        wb_obj = openpyxl.load_workbook(excel_path)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        cell_list = []
        for i in range(1, max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=2)
            cell_obj2 = sheet_obj.cell(row=i, column=5)
            if cell_obj.value is not None:
                cell_dict = {}
                cell_dict['name'] = cell_obj.value
                cell_dict['count'] = cell_obj2.value
                cell_dict['number'] = i
                cell_list.append(cell_dict)
        return cell_list
    return None


def edit_excel_data(path, datas):
    excel_path = join_path(PRODUCT_EXCEL_PATH, path)
    wb_obj = openpyxl.load_workbook(excel_path)
    sheet2 = wb_obj.active
    for data in datas:
        sheet2[f'B{data["number"]}'] = data['name']
        sheet2[f'E{data["number"]}'] = data['count']
    wb_obj.save(excel_path)
    return "Success"


# print(str(datetime.now())[:10])
# def get_contract

def get_excel_vizit(path, all_name, village):
    data = list()
    EXCEL_PATH_VIZIT = join_path(BASE_URL, 'excel_utils', 'excel', 'vizit')
    excel_path = join_path(EXCEL_PATH_VIZIT, path)
    wb_obj = openpyxl.load_workbook(excel_path)
    sheet = wb_obj.active
    max_row = sheet.max_row
    for i in range(1, max_row + 1):
        data_income = {}
        villages = sheet.cell(row=i, column=4).value
        first_last_name = str(sheet.cell(row=i, column=2).value).replace(" ", "")
        all_name = str(all_name).replace(" ", "")
        time = str(sheet.cell(row=i, column=1).value)[:10]
        now_time = str(datetime.now())[:10]
        if villages.__eq__(village) and first_last_name.__eq__(all_name) and time.__eq__(now_time):
            data_income['village'] = villages
            data_income['all_name'] = all_name
            data_income['doctor_name'] = sheet.cell(row=i, column=7).value
            data_income['doctor_category'] = sheet.cell(row=i, column=8).value
            data_income['doctor_type'] = sheet.cell(row=i, column=9).value
            data_income['doctor_phone'] = sheet.cell(row=i, column=10).value
            data_income['comment'] = sheet.cell(row=i, column=11).value
            data.append(data_income)
    return data


def get_excel_income(path, all_name, village):
    data = list()
    excel_path = join_path(INCOME_EXCEL_PATH, path)
    try:
        wb_obj = openpyxl.load_workbook(excel_path)
        sheet = wb_obj.active
        max_row = sheet.max_row
        if village.__eq__("Toshkent Shaxar") or village.__eq__("Toshkent Vil"):
            village = "Ташкент"
        if village.__eq__("Samarqand"):
            village = "Самарканд"
        if village.__eq__("Andijon"):
            village = "Андижан"
        if village.__eq__("Buxoro"):
            village = "Бухара"
        if village.__eq__("Farg'ona"):
            village = "Фергана"
        if village.__eq__("Jizzax"):
            village = "Джизак"
        if village.__eq__("Qoraqalpog'iston"):
            village = "Каракалпакстан"
        if village.__eq__("Qashqadaryo"):
            village = "Кашкадаря"
        if village.__eq__("Navoiy"):
            village = "Навои"
        if village.__eq__("Namangan"):
            village = "Наманган"
        if village.__eq__("Surxondaryo"):
            village = "Сурхандарья"
        if village.__eq__("Sirdaryo"):
            village = "Сырьдарья"
        if village.__eq__("Xorazm"):
            village = "Хорезм"
        for i in range(1, max_row + 1):
            data_income = {}
            villages = sheet.cell(row=i, column=3).value
            first_last_name = str(sheet.cell(row=i, column=4).value).replace(" ", "")
            if villages.__eq__(village) and translate_cyrillic_or_latin(first_last_name, 'cyr').__eq__(
                    translate_cyrillic_or_latin(all_name, 'cyr')):
                data_income['village'] = villages
                data_income['all_name'] = all_name
                data_income['price'] = sheet.cell(row=i, column=5).value
                data_income['company_name'] = sheet.cell(row=i, column=2).value
                data.append(data_income)
        return data
    except Exception:
        return []


def get_excel_data():
    from openpyxl import load_workbook
    import pandas as pd
    excel_path = join_path(DEBIT_EXCEL_PATH, 'Дебеторка 05.05.2023 (3).xlsx')
    wb = openpyxl.load_workbook(excel_path).active
    # wb = load_workbook(filename=excel_path, ).active
    print(wb.values)
    # sheet_names = wb.get_sheet_names()[0]
    # sheet_ranges = wb[name]
    df = pd.DataFrame(list(wb.values)[1:], columns=list(wb.values)[0])
    print(df.head())


def get_excel_debit(path, all_name, village):
    data = list()
    excel_path = join_path(DEBIT_EXCEL_PATH, path)
    try:
        wb_obj = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb_obj.active
        max_row = sheet.max_row
        new_name = all_name.replace(" ", "")
        for i in range(1, max_row + 1):
            data_income = {}
            villages = sheet.cell(row=i, column=5).value
            first_last_name = str(sheet.cell(row=i, column=9).value).replace(" ", "")
            if village.__eq__("Toshkent Shaxar") or village.__eq__("Toshkent Vil"):
                village = "Ташкент"
            if village.__eq__("Samarqand"):
                village = "Самарканд"
            if village.__eq__("Andijon"):
                village = "Андижан"
            if village.__eq__("Buxoro"):
                village = "Бухара"
            if village.__eq__("Farg'ona"):
                village = "Фергана"
            if village.__eq__("Jizzax"):
                village = "Джизак"
            if village.__eq__("Qoraqalpog'iston"):
                village = "Каракалпакстан"
            if village.__eq__("Qashqadaryo"):
                village = "Кашкадаря"
            if village.__eq__("Navoiy"):
                village = "Навои"
            if village.__eq__("Namangan"):
                village = "Наманган"
            if village.__eq__("Surxondaryo"):
                village = "Сурхандарья"
            if village.__eq__("Sirdaryo"):
                village = "Сырьдарья"
            if village.__eq__("Xorazm"):
                village = "Хорезм"
            if villages.__eq__(village) and translate_cyrillic_or_latin(new_name, 'lat').__eq__(
                    translate_cyrillic_or_latin(first_last_name, 'lat')):
                data_income['village'] = village
                data_income['all_name'] = all_name
                price = sheet.cell(row=i, column=12).value
                if price and len(str(price)) != 0 and price != 0:
                    data_income['price'] = sheet.cell(row=i, column=12).value
                else:
                    data_income['price'] = 0
                new_price = sheet.cell(row=i, column=15).value
                if new_price and len(str(new_price).replace(" ", "")) != 0 and new_price != 0:
                    data_income['new_price'] = sheet.cell(row=i, column=15).value
                else:
                    data_income['new_price'] = 0
                received_period = sheet.cell(row=i, column=10).value
                if received_period and len(str(received_period)) != 0:
                    data_income['received_period'] = sheet.cell(row=i, column=10).value
                else:
                    data_income['received_period'] = "No'malum"
                deadline = sheet.cell(row=i, column=11).value
                if deadline and len(str(deadline).replace(" ", "")) != 0:
                    data_income['deadline'] = sheet.cell(row=i, column=11).value
                else:
                    data_income['deadline'] = "No'malum"
                day = sheet.cell(row=i, column=13).value
                if day and len(str(day)) != 0:
                    data_income['day'] = sheet.cell(row=i, column=13).value
                else:
                    data_income['day'] = "No'malum"
                status = sheet.cell(row=i, column=14).value
                if status and len(str(status).replace(" ", "")) != 0:
                    data_income['status'] = sheet.cell(row=i, column=14).value
                else:
                    data_income['status'] = "No'malum"
                data_income['company_name'] = sheet.cell(row=i, column=2).value
                data.append(data_income)
        return data
    except Exception:
        return []
